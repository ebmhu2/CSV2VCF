import sys
import os
import xlrd
import openpyxl
import unicodecsv
from tkinter import messagebox
def xlsToCsvConvertor (xls_filename):
    # Extract the filename along with path without extension.
    csv_filename = xls_filename.rsplit('/', 1)[-1].rsplit('.', 1)[0]

    try:
        # It will load the workbook.
        wb = xlrd.open_workbook(xls_filename)

        # Check the number of sheets in the workbook.
        sh = wb.nsheets
        print("Number of sheets {}".format(sh))
        shName = wb.sheet_names()

        # Loop through the all the sheets.
        for sheet_number in range(sh):
            try:
                # Open the sheet by index.
                wsh = wb.sheet_by_index(sheet_number)

                # Filename to generate the CSV file.
                fileName = xls_filename.rsplit('/', 1)[0] + "/" + csv_filename + "-" + shName[sheet_number] + ".csv"

                # Open the csv file in binary write mode.
                fh = open(fileName, "wb")
                # Uses unicodecsv, so it will handle Unicode characters.
                csv_out = unicodecsv.writer(fh, encoding='utf-8')

                # Loop through the rows of the sheet and write to csv file.
                for row_number in range(wsh.nrows):
                    csv_out.writerow(wsh.row_values(row_number))

                # Close the csv file.
                fh.close()
                messagebox.showinfo("CSV file created successfully")

            except:
                messagebox.showwarning("ERROR",
                                       "Error creating CSV file, because of the following error:\n" + sys.exc_info())

    except:
        messagebox.showwarning("ERROR",
                               "Error opening the file., because of the following error:\n" + sys.exc_info())

def xlsxToCsvHandler(xls_filename):
    # Extract the filename along with path without extension.
    csv_filename = xls_filename.rsplit('/', 1)[-1].rsplit('.', 1)[0]
    xlsFilePath = xls_filename.rsplit('/', 1)[0].rsplit('.', 1)[0]
    print(csv_filename)
    # It will open the workbook.
    try:
        wb = openpyxl.load_workbook(xls_filename)

        # Check the number of sheets in the workbook.
        sh = len(wb.sheetnames)

        print("Number of sheets {}".format(sh))

        shName = wb.sheetnames

        # Loop through the all the sheets.
        for sheet_number in range(sh):
            try:
                # Open the sheet by name.
                wsh = wb[shName[sheet_number]]

                # Filename to generate the CSV file.
                fileName =  xlsFilePath + "/"+ csv_filename + ".csv"
                print(fileName)
                # Open the csv file in binary write mode.
                fh = open(fileName, "wb")

                # Uses unicodecsv, so it will handle Unicode characters.
                csv_out = unicodecsv.writer(fh, encoding='utf-8')

                # Loop through the rows of the sheet and write to csv file.
                for row in wsh.rows:
                    csv_out.writerow(cell.value for cell in row)

                # Close the csv file.
                fh.close()

                messagebox.showinfo("Successful","CSV file created successfully")
            except:
                messagebox.showwarning("ERROR",
                                       "Error creating CSV file, because of the following error:\n" + sys.exc_info())
    except:
        messagebox.showwarning("ERROR",
                               "Error opening the file., because of the following error:\n" + sys.exc_info())

#if __name__ == '__main__':
    #print(sys.argv[1].rsplit('/', 1)[-1].rsplit('.', 1)[-1])
    #if((sys.argv[1].rsplit('/', 1)[-1].rsplit('.', 1)[-1]) == "xls"):
        #xlsToCsvConvertor(sys.argv[1])
    #else:
    #xlsxToCsvHandler("contactNew.xlsx")
